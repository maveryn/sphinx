# sphinx/engine.py
import os, json, random, argparse, re
from dataclasses import asdict
from pathlib import Path
from typing import List, Dict, Any, Tuple, Optional

from .schema import InstanceMeta
from .config import DEFAULT_SEED
from .registry import build_motif_registry, sample_task
from datasets import Dataset, Sequence
from datasets import Image as HFImage
from PIL import Image
from tqdm import tqdm
from openpyxl import Workbook
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment
from concurrent.futures import ProcessPoolExecutor, as_completed

# ------------------------- IO HELPERS ---------------------------------

# NEW: HF-style parquet writer (images as Sequence(Image()))
def save_parquet(rows: List[dict], out_dir: str, filename: str = "dataset.parquet") -> str:
    """
    Save rows to a Parquet file using HuggingFace `datasets`,
    with images stored as Sequence(Image()) (list-of-one).
    Each row in `rows` must contain:
      - 'id' (str), 'image_path' (str), 'problem' (str), 'answer' (str),
        'task' (str), 'sample_seed' (int)
    """
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    path = out / filename

    def gen():
        for r in rows:
            # open the PNG and convert to RGB PIL.Image
            img = Image.open(r["image_path"]).convert("RGB")
            yield {
                "id": str(r["id"]),
                "images": [img],                 # list-of-one PIL.Image
                "problem": str(r["problem"]),
                "answer": str(r["answer"]),      # cast to string
                "task": str(r.get("task", "")),
                "seed": int(r.get("seed", 0)),
                "sample_seed": int(r.get("sample_seed", 0)),
            }

    ds = Dataset.from_generator(gen)
    ds = ds.cast_column("images", Sequence(HFImage()))
    ds.to_parquet(str(path))
    return str(path)


# NEW: Excel writer (image, question, answer only)
def save_excel(rows: List[dict], out_dir: str, filename: str = "metadata.xlsx") -> str:
    """
    Write an Excel file with embedded PNGs in column A and text in B/C.
    Each row should have either 'image_path' or 'image_name', plus 'question' and 'answer'.
    """
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    path = out / filename

    wb = Workbook()
    ws = wb.active
    ws.title = "data"
    ws.append(["image", "question", "answer"])
    ws.freeze_panes = "A2"

    # Column widths
    ws.column_dimensions["A"].width = 60  # image
    ws.column_dimensions["B"].width = 60  # question
    ws.column_dimensions["C"].width = 20  # answer

    for r_idx, r in enumerate(rows, start=2):
        # Text cells
        q = str(r.get("question", "")) if r.get("question", "") is not None else ""
        a = str(r.get("answer", "")) if r.get("answer", "") is not None else ""
        ws.cell(row=r_idx, column=2, value=q).alignment = Alignment(wrap_text=True, vertical="top")
        ws.cell(row=r_idx, column=3, value=a).alignment = Alignment(wrap_text=True, vertical="top")

        # Resolve image path
        img_path = r.get("image_path")
        if not img_path:
            name = r.get("image_name")
            if name:
                img_path = str(out / name)

        # Try embedding; fallback to showing the path as text
        try:
            xl_img = XLImage(str(img_path))
            target_w = 320  # px
            if xl_img.width and xl_img.width > target_w:
                scale = target_w / xl_img.width
                xl_img.width = target_w
                if xl_img.height:
                    xl_img.height = xl_img.height * scale
            ws.add_image(xl_img, f"A{r_idx}")
            if getattr(xl_img, "height", None):
                # row height is in points; use a gentle scaling factor
                ws.row_dimensions[r_idx].height = max(ws.row_dimensions[r_idx].height or 0, xl_img.height * 0.75)
        except Exception:
            ws.cell(row=r_idx, column=1, value=str(img_path)).alignment = Alignment(wrap_text=True, vertical="top")

    wb.save(str(path))
    return str(path)


def _spec_to_dict(s):
    if isinstance(s, dict):
        return {k: _spec_to_dict(v) for k, v in s.items()}
    if hasattr(s, "to_dict"):
        return s.to_dict()
    if hasattr(s, "__dict__"):
        return {k: _spec_to_dict(v) for k, v in s.__dict__.items()}
    return s


def save_metadata(rows: List[InstanceMeta], out_dir: str) -> str:
    out = Path(out_dir)
    out.mkdir(parents=True, exist_ok=True)
    p = out / "metadata.jsonl"
    with p.open("w", encoding="utf-8") as f:
        for r in rows:
            f.write(json.dumps(r.__dict__, ensure_ascii=False) + "\n")
    return str(p)

# ------------------------- RESUME HELPERS -----------------------------

_ID_RE = re.compile(r"sample_(\d{4})$")  # for 'id' fields
_PNG_RE = re.compile(r"sample_(\d{4})\.png$")
_JSON_RE = re.compile(r"sample_(\d{4})\.json$")


def _idx_from_id(sample_id: str) -> Optional[int]:
    m = _ID_RE.search(sample_id)
    return int(m.group(1)) if m else None


def _collect_existing_indices(images_dir: str, metadata_dir: str) -> Tuple[set, set, set]:
    """Return (png_indices, json_indices, all_indices) seen under images/ & metadata/."""
    png_idx, json_idx = set(), set()
    for p in Path(images_dir).glob("sample_*.png"):
        m = _PNG_RE.match(p.name)
        if m:
            png_idx.add(int(m.group(1)))
    for p in Path(metadata_dir).glob("sample_*.json"):
        m = _JSON_RE.match(p.name)
        if m:
            json_idx.add(int(m.group(1)))
    return png_idx, json_idx, png_idx | json_idx


def _load_all_metadata_records(metadata_dir: str) -> List[dict]:
    """Read every per-sample JSON in metadata_dir and return loaded dicts."""
    out: List[dict] = []
    for p in sorted(Path(metadata_dir).glob("sample_*.json")):
        try:
            with open(p, "r", encoding="utf-8") as f:
                out.append(json.load(f))
        except Exception:
            # Skip corrupt records but keep going
            continue
    return out

# ------------------------- PARALLEL WORKER -----------------------------

def _stable_subseed(base_seed: int, index: int) -> int:
    """Deterministic per-sample seed mixer (64-bit) independent of Python's hash randomization."""
    return (int(base_seed) + (index + 1) * 0x9E3779B97F4A7C15) & ((1 << 63) - 1)


def _generate_one_worker(args: Tuple[int, int, str, str]) -> Dict[str, Any]:
    """
    Worker that generates a single sample image + metadata and returns lightweight dicts.
    Args: (i, base_seed, images_dir, metadata_dir)
    """
    i, base_seed, images_dir, metadata_dir = args
    sample_seed = _stable_subseed(base_seed, i)
    rng = random.Random(sample_seed)

    motif_impls = build_motif_registry()
    _, task = sample_task(rng)

    composite, cell_specs, meta = task.generate_instance(motif_impls, rng)
    if composite is None or not meta.get("composite_ready", False):
        raise RuntimeError(f"Task '{getattr(task, 'name', '?')}' did not return a composed image.")

    os.makedirs(images_dir, exist_ok=True)
    os.makedirs(metadata_dir, exist_ok=True)

    fname = f"sample_{i:04d}.png"
    fpath = os.path.join(images_dir, fname)
    composite.save(fpath)

    # Complexity metrics supplied via the task metadata (if available).
    raw_complexity = meta.get("complexity")
    if isinstance(raw_complexity, dict) and raw_complexity:
        complexity_dict = dict(raw_complexity)
    else:
        complexity_score = float(meta.get("complexity_score", 0.5))
        complexity_level = str(meta.get("complexity_level", "MEDIUM"))
        complexity_version = meta.get("complexity_version", "task-meta-v1")
        complexity_dict = {
            "score": complexity_score,
            "level": complexity_level,
            "version": complexity_version,
        }

    task_params = dict(meta)
    if cell_specs:
        task_params["cell_specs"] = [_spec_to_dict(s) for s in cell_specs]

    question = meta.get("question", "")
    raw_answer = meta.get("answer", "")
    answer = "" if raw_answer is None else str(raw_answer)

    meta_obj = InstanceMeta(
        id=f"sample_{i:04d}",
        image_name=fname,
        task=task.name,
        task_params=task_params,
        complexity=complexity_dict,
        seed=base_seed,
        sample_seed=int(sample_seed),
    )

    metadata_fname = f"{meta_obj.id}.json"
    metadata_path = os.path.join(metadata_dir, metadata_fname)
    metadata_payload = asdict(meta_obj)
    metadata_payload.update({
        "question": question,
        "answer": answer,
        "image_path": fpath,
    })
    with open(metadata_path, "w", encoding="utf-8") as meta_file:
        json.dump(metadata_payload, meta_file, ensure_ascii=False, indent=2)

    return {
        "i": i,
        "id": meta_obj.id,
        "fname": fname,
        "fpath": fpath,
        "task": task.name,
        "question": question,
        "answer": answer,
        "task_params": task_params,
        "complexity": complexity_dict,
        "sample_seed": int(sample_seed),
        "metadata_path": metadata_path,
    }


# ------------------------- TOP-LEVEL DRIVER ---------------------------

def generate_dataset(
    n: int = 20,
    out_dir: str = "dataset",
    seed: int = DEFAULT_SEED,
    workers: int = None,
    save_parquet_file: bool = False,
    save_excel_file: bool = False,
    resume: bool = False,
):
    """
    Parallel dataset generation across CPU cores.

    Behavior:
    - When resume=False (default): generate indices [0, n-1], overwriting any existing per-sample files.
    - When resume=True: DO NOT overwrite existing per-sample PNG/JSON files. Only generate the
      missing indices from [0, n-1]. Aggregated outputs (metadata.jsonl / parquet / Excel) are
      rebuilt from all per-sample JSON files found under out_dir/metadata.

    Notes:
    - If only one of PNG/JSON exists for an index, resume mode *skips* that index (no overwrite).
      Remove the stray file or run without --resume to regenerate a clean pair.
    """
    out_root = Path(out_dir)
    out_root.mkdir(parents=True, exist_ok=True)
    images_dir = str(out_root / "images")
    metadata_dir = str(out_root / "metadata")
    os.makedirs(images_dir, exist_ok=True)
    os.makedirs(metadata_dir, exist_ok=True)

    # Determine indices to generate
    if resume:
        png_idx, json_idx, seen_idx = _collect_existing_indices(images_dir, metadata_dir)
        # Skip anything that already has *either* file (avoid overwriting)
        indices_to_generate = [i for i in range(n) if i not in seen_idx]
        found = len(seen_idx)
        msg_prefix = f"Resuming: found {found} existing indices; generating {len(indices_to_generate)} new"
    else:
        indices_to_generate = list(range(n))
        msg_prefix = f"Fresh run: generating {len(indices_to_generate)} samples"

    # Configure workers sensibly
    if workers is None or workers <= 0:
        workers = os.cpu_count() or 1
    if indices_to_generate:
        workers = min(workers, len(indices_to_generate))

    results: List[Dict[str, Any]] = []

    # Fan out to processes (or run inline)
    if indices_to_generate:
        if workers == 1:
            iterator = tqdm(indices_to_generate, desc="Generating dataset", unit="sample")
            for i in iterator:
                results.append(_generate_one_worker((i, seed, images_dir, metadata_dir)))
        else:
            with ProcessPoolExecutor(max_workers=workers) as ex:
                futures = [ex.submit(_generate_one_worker, (i, seed, images_dir, metadata_dir))
                           for i in indices_to_generate]
                for fut in tqdm(as_completed(futures), total=len(indices_to_generate),
                                desc=f"Generating dataset (x{workers})", unit="sample"):
                    results.append(fut.result())

    # Re-read *all* per-sample metadata from disk (existing + newly generated),
    # then rebuild the aggregated artifacts deterministically by index.
    all_records = _load_all_metadata_records(metadata_dir)

    def _safe_idx(rec: dict) -> int:
        idx = _idx_from_id(str(rec.get("id", "")))
        return -1 if idx is None else idx

    all_records = [r for r in all_records if _safe_idx(r) >= 0]
    all_records.sort(key=_safe_idx)

    # Assemble rows for JSONL / Parquet / Excel
    meta_rows: List[InstanceMeta] = []
    parquet_rows: Optional[List[dict]] = [] if save_parquet_file else None
    excel_rows: Optional[List[dict]] = [] if save_excel_file else None

    for r in all_records:
        # JSONL metadata (InstanceMeta)
        try:
            meta_rows.append(InstanceMeta(
                id=r["id"],
                image_name=r["image_name"],
                task=r["task"],
                task_params=r.get("task_params", {}),
                complexity=r.get("complexity", {}),
                seed=int(r.get("seed", seed)),
                sample_seed=int(r.get("sample_seed", 0)),
            ))
        except KeyError:
            # Skip ill-formed records gracefully
            continue

        # Parquet rows
        if parquet_rows is not None:
            parquet_rows.append({
                "id": r["id"],
                "image_path": r.get("image_path", str(out_root / "images" / f"{r['id']}.png")),
                "problem": r.get("question", ""),
                "answer": r.get("answer", ""),
                "task": r.get("task", ""),
                "seed": int(r.get("seed", seed)),
                "sample_seed": int(r.get("sample_seed", 0)),
            })

        # Excel rows
        if excel_rows is not None:
            excel_rows.append({
                "image_path": r.get("image_path", str(out_root / "images" / f"{r['id']}.png")),
                "question": r.get("question", ""),
                "answer": r.get("answer", ""),
            })

    # Write JSONL, Parquet, and Excel
    save_path_jsonl = save_metadata(meta_rows, metadata_dir)
    save_path_parquet = save_parquet(parquet_rows, out_root, filename="dataset.parquet") if parquet_rows is not None else None
    save_path_excel = save_excel(excel_rows, out_root, filename="metadata.xlsx") if excel_rows is not None else None

    # Summary
    total_now = len(all_records)
    parts = [f"{msg_prefix}. Now have {total_now} total samples in '{images_dir}/'."]
    parts.append(f"metadata at {save_path_jsonl}")
    if save_path_parquet:
        parts.append(f"parquet at {save_path_parquet}")
    if save_path_excel:
        parts.append(f"Excel at {save_path_excel}")
    print(" ".join(parts))


if __name__ == "__main__":
    ap = argparse.ArgumentParser()
    ap.add_argument("--n", type=int, default=20)
    ap.add_argument("--out", type=str, default="dataset")
    ap.add_argument("--seed", type=int, default=DEFAULT_SEED)
    ap.add_argument("--workers", type=int, default=max(1, (os.cpu_count() or 1)),
                    help="Number of parallel worker processes (use 1 to disable parallelism).")
    ap.add_argument(
        "--save-parquet",
        action="store_true",
        help="Write HuggingFace-compatible parquet alongside metadata.jsonl.",
    )
    ap.add_argument(
        "--save-excel",
        action="store_true",
        help="Write metadata.xlsx with embedded images alongside metadata.jsonl.",
    )
    ap.add_argument(
        "--resume",
        action="store_true",
        help="Resume without overwriting existing per-sample PNG/JSON; regenerate only missing indices and rebuild aggregated outputs.",
    )
    args = ap.parse_args()
    generate_dataset(
        n=args.n,
        out_dir=args.out,
        seed=args.seed,
        workers=args.workers,
        save_parquet_file=args.save_parquet,
        save_excel_file=args.save_excel,
        resume=args.resume,
    )
