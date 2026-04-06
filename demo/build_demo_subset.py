from __future__ import annotations

import argparse
import json
import random
import re
from collections import defaultdict
from pathlib import Path

import pandas as pd
from huggingface_hub import hf_hub_download
from openpyxl import load_workbook


TASK_DISPLAY = {
    "charts_match_proportions": "Chart Comparison",
    "charts_pie": "Pie Chart",
    "shape_count": "Shape Counting",
    "geometric_position": "Positional Count",
    "geometric_sort": "Shape Sorting",
    "geometric_stack_count": "Stack Count",
    "rect_venn": "Venn Diagram",
    "sequence_arithmetic": "Sequence Arithmetic",
    "sequence_multi_column_arithmetic": "Sequence Multi-Column Arithmetic",
    "sequence_rotation": "Sequence Rotation",
    "symmetry_frieze_groups": "Frieze Groups",
    "symmetry_grid_mirror_fill": "Symmetry Fill",
    "symmetry_scene_mirror_identify": "Mirror Identification",
    "symmetry_wallpaper_groups": "Wallpaper Groups",
    "tiles_decompose_compose": "Tiles Composition",
    "tiles_connected_component": "Tiles Connected Component",
    "tiles_geometry": "Tiles Geometry",
    "tiles_line_intersections": "Tiles Line Intersections",
    "tiles_line_length": "Tiles Line Length",
    "tiles_missing_tiles": "Missing Tiles",
    "tiles_recoloring": "Tiles Recoloring",
    "tiles_shortest_path": "Tiles Shortest Path",
    "transform_pair_infer": "Transform Pair Infer",
    "transform_result_identify": "Transform Result Identify",
    "transform_similarity_identify": "Transform Similarity Identify",
}

HF_DATASET = "xashru/sphinx"
HF_EVAL_FILE = "data/eval-00000-of-00001.parquet"
SAMPLES_PER_TASK = 8
SAMPLE_SEED = 20260405


def _repo_root() -> Path:
    return Path(__file__).resolve().parents[1]


def _default_benchmark_dir() -> Path | None:
    candidates = [
        _repo_root() / "benchmark",
        _repo_root().parent / "source" / "benchmark",
    ]
    return next((path for path in candidates if path.is_dir()), None)


def _resolve_benchmark_files(benchmark_dir: Path | None = None) -> dict[str, Path]:
    benchmark_dir = benchmark_dir or _default_benchmark_dir()
    if benchmark_dir is None:
        raise FileNotFoundError(
            "Could not find benchmark prediction files. Pass --benchmark-dir pointing to the directory "
            "that contains gpt-5_mmr_gpt5score.xlsx and gpt-5-mini_mmr_gpt5score.xlsx."
        )

    files = {
        "gpt-5": benchmark_dir / "gpt-5_mmr_gpt5score.xlsx",
        "gpt-5-mini": benchmark_dir / "gpt-5-mini_mmr_gpt5score.xlsx",
    }
    missing = [str(path) for path in files.values() if not path.is_file()]
    if missing:
        raise FileNotFoundError(
            "Missing benchmark prediction files:\n- " + "\n- ".join(missing)
        )
    return files


def _normalize_problem(text: str) -> str:
    text = text.replace("<image>", " ")
    text = re.sub(r"\s+", " ", text).strip()
    return text


def _normalize_benchmark_question(text: str) -> str:
    text = re.sub(r"^Hint:.*?Question:\s*", "", text, flags=re.S)
    return _normalize_problem(text)


def _hf_parquet_path() -> Path:
    return Path(
        hf_hub_download(
            repo_id=HF_DATASET,
            repo_type="dataset",
            filename=HF_EVAL_FILE,
        )
    )


def _load_hf_buckets() -> dict[tuple[str, str, str], list[dict]]:
    df = pd.read_parquet(_hf_parquet_path())
    buckets: dict[tuple[str, str, str], list[dict]] = defaultdict(list)
    for row in df.itertuples():
        key = (str(row.task).strip(), _normalize_problem(row.problem), str(row.answer).strip())
        image_item = row.images[0]
        buckets[key].append(
            {
                "task_display": str(row.task).strip(),
                "problem": str(row.problem),
                "answer": str(row.answer).strip(),
                "image_bytes": image_item["bytes"],
            }
        )
    return buckets


def _load_benchmark_rows(path: Path) -> dict[tuple[str, str, str], list[dict]]:
    wb = load_workbook(path, read_only=True, data_only=True)
    ws = wb[wb.sheetnames[0]]
    rows = ws.iter_rows(values_only=True)
    headers = next(rows)
    idx = {header: i for i, header in enumerate(headers)}

    buckets: dict[tuple[str, str, str], list[dict]] = defaultdict(list)
    for row in rows:
        task_slug = row[idx["task"]]
        task_display = TASK_DISPLAY[task_slug]
        key = (
            task_display,
            _normalize_benchmark_question(str(row[idx["question"]])),
            str(row[idx["answer"]]).strip(),
        )
        buckets[key].append(
            {
                "id": str(row[idx["id"]]),
                "task_slug": task_slug,
                "task_display": task_display,
                "question": str(row[idx["question"]]),
                "answer": str(row[idx["answer"]]).strip(),
                "prediction": str(row[idx["prediction"]]),
                "normalized_prediction": str(row[idx["normalized_prediction"]]).strip(),
                "correct": bool(row[idx["correct"]]),
                "final_score": bool(row[idx.get("final-score", idx["correct"])]),
            }
        )
    return buckets


def _join_examples(benchmark_dir: Path | None = None) -> list[dict]:
    hf_buckets = _load_hf_buckets()
    benchmark_files = _resolve_benchmark_files(benchmark_dir)
    bench_buckets = {model: _load_benchmark_rows(path) for model, path in benchmark_files.items()}
    examples: list[dict] = []

    all_keys = set(hf_buckets)
    for model_buckets in bench_buckets.values():
        all_keys &= set(model_buckets)

    for key in sorted(all_keys):
        hf_rows = hf_buckets[key]
        model_rows = {model: buckets[key] for model, buckets in bench_buckets.items()}
        expected = len(hf_rows)
        if any(len(rows) != expected for rows in model_rows.values()):
            raise ValueError(f"Mismatched counts for key {key!r}")

        for idx in range(expected):
            gpt5 = model_rows["gpt-5"][idx]
            mini = model_rows["gpt-5-mini"][idx]
            if gpt5["id"] != mini["id"] or gpt5["task_slug"] != mini["task_slug"]:
                raise ValueError(f"Model row mismatch for key {key!r} at index {idx}")

            hf_row = hf_rows[idx]
            sample_id = f"{gpt5['task_slug']}__{gpt5['id']}"
            examples.append(
                {
                    "sample_id": sample_id,
                    "raw_id": gpt5["id"],
                    "task_slug": gpt5["task_slug"],
                    "task_display": gpt5["task_display"],
                    "problem": _normalize_problem(hf_row["problem"]),
                    "answer": hf_row["answer"],
                    "image_bytes": hf_row["image_bytes"],
                    "predictions": {
                        "gpt-5": gpt5,
                        "gpt-5-mini": mini,
                    },
                }
            )

    return examples


def _write_subset(examples: list[dict], out_dir: Path) -> None:
    rng = random.Random(SAMPLE_SEED)
    grouped: dict[str, list[dict]] = defaultdict(list)
    for example in examples:
        grouped[example["task_slug"]].append(example)

    selected: list[dict] = []
    for task_slug in sorted(TASK_DISPLAY):
        task_examples = grouped[task_slug]
        if len(task_examples) < SAMPLES_PER_TASK:
            raise ValueError(f"Task {task_slug} has only {len(task_examples)} joinable rows")
        picks = rng.sample(task_examples, SAMPLES_PER_TASK)
        picks.sort(key=lambda row: row["sample_id"])
        selected.extend(picks)

    images_dir = out_dir / "images"
    images_dir.mkdir(parents=True, exist_ok=True)
    examples_path = out_dir / "examples.jsonl"
    summary_path = out_dir / "summary.json"

    task_counts: dict[str, int] = defaultdict(int)
    model_scores: dict[str, dict[str, int]] = defaultdict(lambda: {"correct": 0, "total": 0})
    with examples_path.open("w", encoding="utf-8") as fh:
        for example in selected:
            task_slug = example["task_slug"]
            task_counts[task_slug] += 1
            image_rel = Path("images") / task_slug / f"{example['sample_id']}.png"
            image_path = out_dir / image_rel
            image_path.parent.mkdir(parents=True, exist_ok=True)
            image_path.write_bytes(example["image_bytes"])

            payload = {
                "id": example["sample_id"],
                "title": example["task_display"],
                "task": task_slug,
                "mode": "image_qa",
                "prompt": example["problem"],
                "media": [
                    {
                        "type": "image",
                        "src": image_rel.as_posix(),
                        "alt": example["task_display"],
                    }
                ],
                "answer": {
                    "canonical": example["answer"],
                },
                "predictions": [
                    {
                        "model": "GPT-5",
                        "answer": example["predictions"]["gpt-5"]["prediction"],
                        "correct": bool(example["predictions"]["gpt-5"]["final_score"]),
                    },
                    {
                        "model": "GPT-5 Mini",
                        "answer": example["predictions"]["gpt-5-mini"]["prediction"],
                        "correct": bool(example["predictions"]["gpt-5-mini"]["final_score"]),
                    },
                ],
                "meta": {
                    "raw_id": example["raw_id"],
                    "task_display": example["task_display"],
                    "source_dataset": HF_DATASET,
                    "source_split": "eval",
                    "models": {
                        "gpt-5": {
                            "normalized_prediction": example["predictions"]["gpt-5"]["normalized_prediction"],
                            "correct": bool(example["predictions"]["gpt-5"]["final_score"]),
                        },
                        "gpt-5-mini": {
                            "normalized_prediction": example["predictions"]["gpt-5-mini"]["normalized_prediction"],
                            "correct": bool(example["predictions"]["gpt-5-mini"]["final_score"]),
                        },
                    },
                },
            }
            for prediction in payload["predictions"]:
                stats = model_scores[prediction["model"]]
                stats["total"] += 1
                if prediction["correct"]:
                    stats["correct"] += 1
            fh.write(json.dumps(payload, ensure_ascii=False) + "\n")

    model_summary = {}
    for model, stats in sorted(model_scores.items()):
        accuracy = stats["correct"] / stats["total"] if stats["total"] else 0.0
        model_summary[model] = {
            "correct": stats["correct"],
            "total": stats["total"],
            "accuracy": round(accuracy, 4),
        }

    default_compare_models = [
        model
        for model, _ in sorted(
            model_summary.items(),
            key=lambda item: (-item[1]["accuracy"], item[0]),
        )[:2]
    ]

    summary = {
        "dataset": HF_DATASET,
        "split": "eval",
        "samples_per_task": SAMPLES_PER_TASK,
        "total_examples": len(selected),
        "tasks": {task: task_counts[task] for task in sorted(task_counts)},
        "models": sorted(model_summary),
        "model_summary": model_summary,
        "default_compare_models": default_compare_models,
    }
    summary_path.write_text(json.dumps(summary, indent=2), encoding="utf-8")


def main() -> None:
    parser = argparse.ArgumentParser(description="Build the published SPHINX demo subset.")
    parser.add_argument(
        "--benchmark-dir",
        type=Path,
        default=None,
        help="Directory containing gpt-5_mmr_gpt5score.xlsx and gpt-5-mini_mmr_gpt5score.xlsx.",
    )
    parser.add_argument(
        "--out",
        type=Path,
        default=Path(__file__).resolve().parent,
        help="Output directory for examples.jsonl, summary.json, and images/.",
    )
    args = parser.parse_args()

    out_dir = args.out
    examples = _join_examples(args.benchmark_dir)
    _write_subset(examples, out_dir)
    print(f"Wrote demo subset to {out_dir}")


if __name__ == "__main__":
    main()
